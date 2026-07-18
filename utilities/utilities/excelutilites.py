from openpyxl import load_workbook

class ExcelUtilities:
    @staticmethod
    def get_row(file,sheet):
        file=load_workbook(file)
        sheet=file[sheet]
        return sheet.max_row

    @staticmethod
    def get_column(file,sheet):
        file=load_workbook(file)
        sheet=file[sheet]
        return sheet.max_column

    @staticmethod
    def read_data(file,sheet,row,column):
        file=load_workbook(file)
        sheet=file[sheet]
        return sheet.cell(row=row,column=column).value

